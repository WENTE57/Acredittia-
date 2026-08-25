"""Tareas del worker: revisión IA, extracciones IA y generación de reportes.

Registro único: importar este módulo es lo que puebla `services.jobs.TAREAS`.
Lo importa `main.py` al arrancar la API y el worker de Celery en su propio
proceso; sin ese import `enqueue()` no encuentra la tarea y solo deja un error
en el log.

Cuatro reglas comunes a todas las tareas:

* **Sesión propia.** No hay request, así que cada tarea abre `worker_session`
  con el `company_id` recibido: el contexto de RLS se fija ahí y no se hereda
  del proceso que encoló.
* **Kwargs serializables.** Celery transporta JSON, así que todo entra como
  string y se convierte a `uuid.UUID` o `date` dentro de la tarea.
* **No propagan excepciones.** Un fallo marca el job `failed` con el mensaje y
  queda en `logging`. El worker nunca ve la excepción: un reintento automático
  sobre una revisión ya aplicada duplicaría alertas y fechas de vencimiento.
* **Idempotencia.** Cada tarea sale sin hacer nada si el job ya está `done`, y
  borra los hallazgos previos antes de reescribirlos, de modo que un reintento
  manual no los acumule.

Las extracciones (`extraer_*`) **no crean entidades**: dejan los valores
detectados en `ia_reviews.campos_extraidos` para que el usuario confirme el
formulario y el endpoint correspondiente cree el registro (§12.1).
"""
from __future__ import annotations

import hashlib
import io
import logging
import os
import random
import uuid
from datetime import date, datetime, timezone

from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..database import worker_session
from ..models import (Alerta, Contrato, Documento, DocumentoArchivo, Faena,
                      IaHallazgo, IaReview, Reporte, Sujeto)
from . import actividad
from .checklist import (calc_estado_doc, calc_estado_sujeto,
                        vencimiento_por_plantilla)
from .ia import get_reviewer
from .jobs import tarea
from .storage import get_storage

logger = logging.getLogger("acredittia.tasks")

# Gravedad relativa de `alerta_severidad`, de más a menos grave. Se usa para
# elegir la severidad de la alerta a partir del peor hallazgo de la revisión.
ORDEN_SEVERIDAD = ("critica", "alta", "media", "baja", "advertencia",
                   "informativa")

# Severidad de la alerta según el código de hallazgo de tipo error (§12). Un
# RUT que no coincide o un sello no reconocido invalidan el documento ante el
# mandante; una resolución baja o una firma ausente se pueden resubir.
SEVERIDAD_POR_CODIGO = {
    "RUT_NO_COINCIDE": "critica",
    "SELLO_NO_RECONOCIDO": "critica",
    "CLASE_LICENCIA_INSUFICIENTE": "critica",
    "DOC_ILEGIBLE": "alta",
    "FIRMA_NO_ENCONTRADA": "alta",
    "VENCIMIENTO_AUSENTE": "media",
}
SEVERIDAD_ERROR_DEFECTO = "alta"

# Recurso de `routers/reportes.py` que alimenta cada tipo de reporte. El enum
# `reporte_tipo` de la BD no tiene un valor por vista, así que el recurso real
# viaja además en `reportes.params['recurso']` y tiene prioridad sobre esta
# tabla (lo escribe /exportaciones al derivar una exportación grande).
RECURSO_POR_TIPO = {
    "estado_acreditacion": "contratos",
    "cumplimiento_requisitos": "requisitos",
    "personal_acreditado": "personal",
    "equipos_vehiculos": "equipos",
    "vencimientos": "documentos",
}


# ============================================================================
# Helpers comunes
# ============================================================================
def contexto_de_documento(doc: Documento) -> str:
    """Valor de `ia_context` que corresponde al dueño del documento.

    El contexto condiciona qué valida la IA, y se deduce de quién cuelga el
    documento: los de empresa cuelgan del contrato, los de personal y equipo de
    un sujeto, y el expediente EMSIPOR se marca con su propia bandera porque
    cuelga de un trabajador pero se valida con otro criterio.
    """
    if doc.es_emsipor:
        return "emsipor"
    if doc.contrato_id:
        return "empresa"
    if doc.sujeto is not None:
        return "equipo" if doc.sujeto.tipo == "equipo" else "personal"
    return "personal"


def _rng(semilla: bytes, sal: str) -> random.Random:
    """Generador determinista: el mismo archivo produce siempre lo mismo."""
    h = hashlib.sha256(semilla[:4096] + sal.encode()).hexdigest()
    return random.Random(int(h[:16], 16))


def _dv_rut(cuerpo: str) -> str:
    """Dígito verificador módulo 11 (mismo algoritmo que `security.validar_rut`)."""
    suma, factor = 0, 2
    for c in reversed(cuerpo):
        suma += int(c) * factor
        factor = 2 if factor == 7 else factor + 1
    resto = 11 - (suma % 11)
    return "0" if resto == 11 else "k" if resto == 10 else str(resto)


def _rut_plausible(n: int) -> str:
    """Formatea 12345678 como '12.345.678-5' con dígito verificador válido.

    El verificador va en mayúscula (`-K`), que es la forma canónica en Chile y
    la que espera ver el usuario en el formulario; `security.validar_rut` acepta
    ambas.
    """
    return f"{n:,}".replace(",", ".") + "-" + _dv_rut(str(n)).upper()


def _peor(severidades: list[str]) -> str:
    """Severidad más grave de la lista, o la de defecto si viene vacía."""
    for s in ORDEN_SEVERIDAD:
        if s in severidades:
            return s
    return SEVERIDAD_ERROR_DEFECTO


def _iso_o_none(v) -> str | None:
    return v.isoformat() if v else None


def _cargar_review(db: Session, review_id: str | uuid.UUID,
                   company_id: uuid.UUID) -> IaReview | None:
    r = db.get(IaReview, uuid.UUID(str(review_id)))
    if r is None or r.company_id != company_id:
        logger.error("ia_review %s inexistente o de otra empresa", review_id)
        return None
    return r


def _marcar_procesando(db: Session, review: IaReview) -> None:
    review.status = "processing"
    review.started_at = datetime.now(timezone.utc)
    db.commit()


def _fallar(db: Session, review_id: uuid.UUID, mensaje: str) -> None:
    """Cierra el job como `failed` sin propagar la excepción al worker.

    Recibe el id y no la instancia: tras el `rollback` los atributos del objeto
    están expirados y recargarlos dentro del manejo de errores es justo donde no
    interesa que aparezca una segunda excepción.
    """
    db.rollback()
    review = db.get(IaReview, review_id)
    if review is None:
        return
    review.status = "failed"
    review.error = mensaje[:500]
    review.finished_at = datetime.now(timezone.utc)
    db.commit()


def _reescribir_hallazgos(db: Session, review: IaReview,
                          hallazgos: list[dict]) -> None:
    """Sustituye los hallazgos del job (idempotente ante reintentos)."""
    db.execute(delete(IaHallazgo).where(IaHallazgo.review_id == review.id))
    for h in hallazgos:
        db.add(IaHallazgo(
            review_id=review.id, tipo=h["tipo"], codigo=h["codigo"],
            mensaje=h["mensaje"], campo=h.get("campo"),
            valor_detectado=h.get("valor_detectado"),
            valor_esperado=h.get("valor_esperado"),
        ))


def _fecha_extraida(campos: dict, clave: str = "fecha_vencimiento") -> date | None:
    """Interpreta una fecha ISO de `campos_extraidos`; None si no es usable."""
    valor = (campos or {}).get(clave)
    if not valor:
        return None
    try:
        return date.fromisoformat(str(valor)[:10])
    except ValueError:
        logger.warning("fecha_vencimiento no interpretable: %r", valor)
        return None


# ============================================================================
# §12 — Revisión de un documento ya subido
# ============================================================================
@tarea("revisar_documento")
def revisar_documento(archivo_id: str, company_id: str,
                      review_id: str | None = None) -> None:
    """Revisa un archivo con la IA y aplica la decisión sobre el documento.

    Reglas de decisión (§12): sin hallazgos de tipo error el documento pasa a
    `ok` y se autocompleta `vence` —con la fecha que extrajo la IA o, si no la
    extrajo, con la vigencia de la plantilla del requisito, porque sin fecha el
    documento quedaría `ok` para siempre—. Con al menos un error el documento
    permanece en `falta` y se emite una `Alerta` de origen `ia` con la severidad
    del hallazgo más grave.

    Lo que hizo la tarea se resume en `campos_extraidos['_accion_aplicada']`,
    que es lo que devuelve `GET /ia/revisiones/{job_id}`: sin ese rastro el
    usuario ve el veredicto de la IA pero no qué se cambió en su expediente.
    """
    cid = uuid.UUID(company_id)
    with worker_session(company_id=cid) as db:
        archivo = db.get(DocumentoArchivo, uuid.UUID(archivo_id))
        if archivo is None or archivo.company_id != cid:
            logger.error("archivo %s inexistente o de otra empresa", archivo_id)
            return
        doc = archivo.documento
        contexto = contexto_de_documento(doc)

        if review_id:
            review = _cargar_review(db, review_id, cid)
            if review is None:
                return
            if review.status == "done":
                logger.info("revisión %s ya aplicada; nada que hacer", review.id)
                return
        else:
            # Revisión disparada sin job previo (reproceso desde el worker): se
            # crea la fila para que el resultado sea consultable por polling.
            review = IaReview(
                company_id=cid, archivo_id=archivo.id, context=contexto,
                status="queued",
                campos_extraidos={"documento_id": str(doc.id),
                                  "blob_path": archivo.blob_path,
                                  "filename": archivo.filename})
            db.add(review)
            db.flush()

        rid = review.id
        base = dict(review.campos_extraidos or {})
        _marcar_procesando(db, review)

        try:
            contenido = get_storage().read(archivo.blob_path)
            r = get_reviewer().revisar(contenido, archivo.filename, doc, contexto)
        except Exception as e:  # noqa: BLE001
            logger.exception("revisión %s falló al leer o analizar el blob", rid)
            _fallar(db, rid, f"{type(e).__name__}: {e}")
            return

        try:
            hallazgos = [{"tipo": h.tipo, "codigo": h.codigo, "mensaje": h.mensaje}
                         for h in r.hallazgos]
            errores = [h for h in hallazgos if h["tipo"] == "error"]

            campos = {**base, **(r.campos_extraidos or {})}
            review.resultado = r.resultado
            review.confianza = round(float(r.confianza), 3)
            review.status = "done"
            review.error = None
            _reescribir_hallazgos(db, review, hallazgos)
            archivo.ia_review_id = review.id

            if errores:
                doc.estado = "falta"
                severidad = _peor([SEVERIDAD_POR_CODIGO.get(h["codigo"],
                                                            SEVERIDAD_ERROR_DEFECTO)
                                   for h in errores])
                titulo = f"Documento observado por IA: {doc.titulo}"
                db.add(Alerta(
                    company_id=cid, severidad=severidad,
                    estado="bloqueante" if severidad == "critica" else "nueva",
                    origen="ia", titulo=titulo,
                    descripcion="; ".join(h["mensaje"] for h in errores),
                    documento_id=doc.id, sujeto_id=doc.sujeto_id,
                    contrato_id=doc.contrato_id,
                ))
                accion = (f"documento mantenido en falta; alerta {severidad} "
                          f"generada ({', '.join(h['codigo'] for h in errores)})")
            else:
                doc.estado = "ok"
                extraida = _fecha_extraida(campos)
                if extraida:
                    doc.vence = extraida
                    origen_fecha = "fecha extraída por la IA"
                elif doc.vence is None:
                    doc.vence = vencimiento_por_plantilla(db, doc)
                    origen_fecha = ("vigencia de la plantilla" if doc.vence
                                    else "sin vencimiento conocido")
                else:
                    origen_fecha = "vencimiento informado previamente"
                etiqueta = ("ok" if r.resultado == "validado"
                            else "ok con observaciones")
                accion = (f"documento marcado {etiqueta}; "
                          f"vence={_iso_o_none(doc.vence) or 'null'} ({origen_fecha})")

            doc.estado_calc = calc_estado_doc(doc)
            if doc.sujeto_id:
                docs = list(db.scalars(select(Documento).where(
                    Documento.company_id == cid,
                    Documento.sujeto_id == doc.sujeto_id)))
                for d in docs:
                    d.estado_calc = calc_estado_doc(d)
                sujeto = db.get(Sujeto, doc.sujeto_id)
                if sujeto is not None:
                    sujeto.estado = calc_estado_sujeto(docs, sujeto.estado)

            campos["_accion_aplicada"] = accion
            review.campos_extraidos = campos          # reasignar: JSONB no rastrea mutaciones
            review.finished_at = datetime.now(timezone.utc)

            # user_id=None → el feed de actividad lo muestra como "Sistema IA".
            actividad.log(db, cid, "alerta_ia", "alertas_ia",
                          f"Revisión IA de '{doc.titulo}': {r.resultado}; {accion}",
                          user_id=None, entidad_tipo="documento",
                          entidad_id=doc.id)
            db.commit()
            logger.info("revisión %s aplicada: %s", rid, accion)
        except Exception as e:  # noqa: BLE001
            logger.exception("revisión %s falló al aplicar la decisión", rid)
            _fallar(db, rid, f"{type(e).__name__}: {e}")


# ============================================================================
# §12.1 — Extracción de sujeto (cédula / padrón)
# ============================================================================
NOMBRES = ["Juan Pérez Soto", "María González Rojas", "Carlos Muñoz Araya",
           "Patricia Silva Contreras", "Luis Fuentes Cárdenas",
           "Rodrigo Tapia Núñez", "Claudia Vergara Pinto",
           "Sergio Alarcón Bustos"]
CARGOS = ["Conductor Nacional", "Operador de Equipo Pesado",
          "Supervisor de Terreno", "Mecánico Mantenedor",
          "Prevencionista de Riesgos", "Ayudante de Operaciones"]
MARCAS = {
    "Mercedes-Benz": ["Actros 2646", "Axor 3344", "Sprinter 515"],
    "Volvo": ["FH 460", "FM 440", "FMX 500"],
    "Scania": ["R 450", "G 410", "P 360"],
    "Toyota": ["Hilux 4x4", "Land Cruiser"],
    "Caterpillar": ["420F", "950GC", "12M3"],
}
TIPOS_EQUIPO_SUGERIDOS = ["Tracto-Camión", "Camión Pluma", "Camión Aljibe",
                          "Camioneta", "Retroexcavadora", "Cama Baja",
                          "Motoniveladora", "Grúa Horquilla"]


def _patente_plausible(rng: random.Random) -> str:
    """Formato chileno moderno de 4 letras + 2 dígitos (BBBB·11)."""
    letras = "BCDFGHJKLPRSTVWXYZ"
    return ("".join(rng.choice(letras) for _ in range(4))
            + f"{rng.randint(10, 99)}")


@tarea("extraer_sujeto")
def extraer_sujeto(review_id: str, company_id: str, blob_path: str,
                   tipo: str) -> None:
    """Pre-llena el formulario de alta de un trabajador o de un equipo.

    Extracción simulada determinista por hash del contenido: el mismo archivo
    devuelve siempre los mismos valores, que es lo que hace usable un prototipo
    (y lo que permite escribir pruebas). El adaptador real sustituye solo el
    cuerpo de esta función; el contrato de `campos_extraidos` no cambia.

    No crea el sujeto: el usuario revisa los campos con confianza baja y
    confirma en `POST /personal` o `POST /equipos` (§12.1).
    """
    cid = uuid.UUID(company_id)
    with worker_session(company_id=cid) as db:
        review = _cargar_review(db, review_id, cid)
        if review is None:
            return
        if review.status == "done":
            return
        rid = review.id
        base = dict(review.campos_extraidos or {})
        _marcar_procesando(db, review)
        try:
            contenido = get_storage().read(blob_path)
            rng = _rng(contenido, f"sujeto:{tipo}")

            if tipo == "cedula":
                campos = {
                    "nombre": rng.choice(NOMBRES),
                    "rut": _rut_plausible(rng.randint(6_000_000, 24_999_999)),
                    "cargo_sugerido": rng.choice(CARGOS),
                }
            else:                                   # padron | permiso de circulación
                marca = rng.choice(list(MARCAS))
                campos = {
                    "patente": _patente_plausible(rng),
                    "marca": marca,
                    "modelo": rng.choice(MARCAS[marca]),
                    "anio": rng.randint(2012, date.today().year),
                    "tipo_equipo_sugerido": rng.choice(TIPOS_EQUIPO_SUGERIDOS),
                }

            # Confianza por campo: el frontend resalta los < 0.85 para revisión
            # manual, así que se informa campo a campo y no solo el global.
            por_campo = {k: round(rng.uniform(0.72, 0.99), 3) for k in campos}
            global_ = round(sum(por_campo.values()) / len(por_campo), 3)
            campos_extraidos = {
                **base, **campos,
                "confianza": global_,
                "confianza_campos": por_campo,
                "_accion_aplicada": ("valores propuestos para el formulario; "
                                     "no se creó ningún registro"),
            }

            review.campos_extraidos = campos_extraidos
            review.confianza = global_
            review.resultado = "validado" if global_ >= 0.85 else "con_observaciones"
            review.status = "done"
            review.finished_at = datetime.now(timezone.utc)
            _reescribir_hallazgos(db, review, [
                {"tipo": "info", "codigo": "EXTRACCION_OK",
                 "mensaje": f"Datos extraídos de {tipo}."},
            ] + [
                {"tipo": "warning", "codigo": "CONFIANZA_BAJA",
                 "mensaje": f"Revise manualmente el campo '{k}'.",
                 "campo": k, "valor_detectado": str(campos[k])}
                for k, v in sorted(por_campo.items()) if v < 0.85
            ])
            db.commit()
        except Exception as e:  # noqa: BLE001
            logger.exception("extracción de sujeto %s falló", review_id)
            _fallar(db, rid, f"{type(e).__name__}: {e}")


# ============================================================================
# §12.1 — Extracción de contrato
# ============================================================================
TIPOS_SERVICIO = ["Transporte de personal", "Transporte de insumos",
                  "Movimiento de tierra", "Mantención mecánica",
                  "Servicios generales", "Obras civiles"]


@tarea("extraer_contrato")
def extraer_contrato(review_id: str, company_id: str, blob_path: str) -> None:
    """Pre-llena el alta de contrato leyendo el PDF o Word firmado.

    La faena se devuelve emparejada contra las `faenas` reales del catálogo
    (`faena_id`), no como texto libre: el formulario necesita el id para el
    selector, y si la detección falla es mejor `faena_id=null` con el nombre
    detectado que un id inventado.
    """
    cid = uuid.UUID(company_id)
    with worker_session(company_id=cid) as db:
        review = _cargar_review(db, review_id, cid)
        if review is None:
            return
        if review.status == "done":
            return
        rid = review.id
        base = dict(review.campos_extraidos or {})
        _marcar_procesando(db, review)
        try:
            contenido = get_storage().read(blob_path)
            rng = _rng(contenido, "contrato")

            faenas = list(db.scalars(select(Faena).where(Faena.activa)
                                     .order_by(Faena.nombre)))
            faena = rng.choice(faenas) if faenas else None
            inicio = date(rng.randint(date.today().year - 1, date.today().year),
                          rng.randint(1, 12), rng.randint(1, 28))
            meses = rng.choice([12, 18, 24, 36])
            termino = date(inicio.year + (inicio.month - 1 + meses) // 12,
                           (inicio.month - 1 + meses) % 12 + 1, inicio.day)

            campos = {
                "mandante": faena.mandante if faena else "Mandante no detectado",
                "faena_id": str(faena.id) if faena else None,
                "faena_nombre": faena.nombre if faena else None,
                "tipo_servicio": rng.choice(TIPOS_SERVICIO),
                "contratista": (db.scalar(select(Contrato.nombre).where(
                    Contrato.company_id == cid).limit(1)) or "Empresa contratista"),
                "fecha_inicio": inicio.isoformat(),
                "fecha_termino": termino.isoformat(),
                "renovacion_automatica": bool(rng.random() < 0.45),
            }
            por_campo = {k: round(rng.uniform(0.70, 0.99), 3) for k in campos}
            if faena is None:
                por_campo["faena_id"] = 0.0
            global_ = round(sum(por_campo.values()) / len(por_campo), 3)

            observaciones: list[dict] = [
                {"tipo": "info", "codigo": "CONTRATO_RECONOCIDO",
                 "mensaje": "Se identificó la carátula del contrato y sus fechas."},
            ]
            if faena is None:
                observaciones.append({
                    "tipo": "warning", "codigo": "FAENA_NO_EMPAREJADA",
                    "mensaje": ("No se pudo asociar la faena detectada con el "
                                "catálogo; selecciónela manualmente."),
                    "campo": "faena_id"})
            else:
                observaciones.append({
                    "tipo": "info", "codigo": "FAENA_EMPAREJADA",
                    "mensaje": f"Faena asociada a '{faena.nombre}'.",
                    "campo": "faena_id", "valor_detectado": faena.nombre})
            if campos["renovacion_automatica"]:
                observaciones.append({
                    "tipo": "warning", "codigo": "RENOVACION_AUTOMATICA",
                    "mensaje": ("El contrato incluye cláusula de renovación "
                                "automática; verifique el plazo de aviso."),
                    "campo": "renovacion_automatica"})
            observaciones += [
                {"tipo": "warning", "codigo": "CONFIANZA_BAJA",
                 "mensaje": f"Revise manualmente el campo '{k}'.", "campo": k}
                for k, v in sorted(por_campo.items()) if v < 0.85
            ]

            review.campos_extraidos = {
                **base, **campos, "confianza": global_,
                "confianza_campos": por_campo,
                "_accion_aplicada": ("datos propuestos para el alta de contrato; "
                                     "no se creó ningún contrato"),
            }
            review.confianza = global_
            review.resultado = "validado" if global_ >= 0.85 else "con_observaciones"
            review.status = "done"
            review.finished_at = datetime.now(timezone.utc)
            _reescribir_hallazgos(db, review, observaciones)
            db.commit()
        except Exception as e:  # noqa: BLE001
            logger.exception("extracción de contrato %s falló", review_id)
            _fallar(db, rid, f"{type(e).__name__}: {e}")


# ============================================================================
# §12.1 — Extracción de carpeta de arranque
# ============================================================================
# Conjunto de respaldo cuando el archivo no es tabular (PDF o Word escaneado).
# Son los requisitos que aparecen en toda carpeta de arranque minera chilena.
ARRANQUE_SIMULADO: dict[str, list[tuple[str, bool]]] = {
    "empresa": [
        ("Certificado de Adhesión a Mutualidad", True),
        ("Reglamento Interno de Orden, Higiene y Seguridad", True),
        ("Programa de Prevención de Riesgos", True),
        ("Certificado F30-1 Dirección del Trabajo", True),
        ("Póliza de Responsabilidad Civil", True),
        ("Matriz de Identificación de Peligros (IPER)", True),
        ("Plan de Emergencia y Evacuación", False),
        ("Certificado de Inscripción en Registro de Contratistas", False),
    ],
    "personal": [
        ("Contrato de Trabajo", True),
        ("Anexo de Traslado a Faena", True),
        ("Examen Ocupacional Vigente", True),
        ("Inducción Hombre Nuevo", True),
        ("Entrega de Elementos de Protección Personal", True),
        ("Certificado de Capacitación Trabajo en Altura Física", False),
        ("Registro ODI (Obligación de Informar)", True),
    ],
    "equipo": [
        ("Permiso de Circulación", True),
        ("Seguro Obligatorio (SOAP)", True),
        ("Revisión Técnica Vigente", True),
        ("Certificado de Mantención Preventiva", True),
        ("Certificación de Sistema de Extinción", False),
        ("Check List de Equipo Operativo", True),
    ],
}

# Palabras que delatan el ámbito de un requisito leído de una planilla.
PISTAS_AMBITO = {
    "personal": ("trabajador", "personal", "conductor", "operador",
                 "examen", "induccion", "capacitacion", "contrato de trabajo",
                 "epp", "odi"),
    "equipo": ("equipo", "vehiculo", "vehículo", "camion", "camión", "patente",
               "circulacion", "circulación", "soap", "revision tecnica",
               "revisión técnica", "mantencion", "mantención"),
}


def _ambito_de_titulo(titulo: str) -> str:
    """Ámbito inferido del texto del requisito; 'empresa' por defecto.

    LIMITACIÓN: es una heurística por palabras clave, no una clasificación
    semántica. Basta para pre-agrupar la propuesta, y el usuario reasigna el
    ámbito al confirmar; nada se crea automáticamente.
    """
    t = titulo.lower()
    for ambito, pistas in PISTAS_AMBITO.items():
        if any(p in t for p in pistas):
            return ambito
    return "empresa"


def _titulos_tabulares(contenido: bytes, filename: str) -> list[str]:
    """Títulos de la primera columna de un CSV o XLSX. Lista vacía si no aplica."""
    ext = os.path.splitext(filename or "")[1].lower()
    titulos: list[str] = []
    if ext == ".csv":
        import csv
        for codec in ("utf-8-sig", "latin-1"):
            try:
                texto = contenido.decode(codec)
                break
            except UnicodeDecodeError:
                continue
        else:
            return []
        # El separador se detecta porque las planillas chilenas usan ';'.
        try:
            dialecto = csv.Sniffer().sniff(texto[:4096], delimiters=";,\t")
        except csv.Error:
            dialecto = csv.excel
        for fila in csv.reader(io.StringIO(texto), dialecto):
            if fila and fila[0].strip():
                titulos.append(fila[0].strip())
    elif ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook       # import perezoso: opcional
        except ImportError:
            logger.info("openpyxl no disponible; se usa el conjunto simulado")
            return []
        try:
            wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        except Exception:                            # .xls antiguo, cifrado, corrupto
            logger.info("planilla ilegible por openpyxl; conjunto simulado")
            return []
        for fila in wb.worksheets[0].iter_rows(values_only=True):
            if fila and fila[0] is not None and str(fila[0]).strip():
                titulos.append(str(fila[0]).strip())
        wb.close()

    # Se descarta la primera fila si parece encabezado y las líneas de relleno.
    limpios = [t for t in titulos if 4 <= len(t) <= 160]
    if limpios and limpios[0].lower() in ("requisito", "documento", "titulo",
                                          "título", "descripcion", "descripción"):
        limpios = limpios[1:]
    return limpios[:200]


@tarea("extraer_carpeta_arranque")
def extraer_carpeta_arranque(review_id: str, company_id: str, contrato_id: str,
                             blob_path: str) -> None:
    """Propone los requisitos de la Carpeta de Arranque agrupados por ámbito.

    Si el archivo es tabular se leen los títulos de la primera columna, que es
    como llegan las carpetas reales (una planilla del mandante); si no lo es, se
    devuelve un conjunto plausible para que el flujo de confirmación se pueda
    probar igual.

    **No crea requisitos**: el usuario confirma la selección con
    `POST /contratos/{id}/requisitos?bulk=true`, que ya omite los duplicados.
    """
    cid = uuid.UUID(company_id)
    with worker_session(company_id=cid) as db:
        review = _cargar_review(db, review_id, cid)
        if review is None:
            return
        if review.status == "done":
            return
        rid = review.id
        base = dict(review.campos_extraidos or {})
        filename = str(base.get("filename") or os.path.basename(blob_path))
        _marcar_procesando(db, review)
        try:
            contenido = get_storage().read(blob_path)
            titulos = _titulos_tabulares(contenido, filename)
            rng = _rng(contenido, "arranque")

            propuesta: dict[str, list[dict]] = {"empresa": [], "personal": [],
                                                "equipo": []}
            if titulos:
                fuente = "planilla"
                vistos: set[str] = set()
                for t in titulos:
                    clave = t.lower()
                    if clave in vistos:
                        continue
                    vistos.add(clave)
                    propuesta[_ambito_de_titulo(t)].append({
                        "titulo": t,
                        "obligatorio": True,
                        "confianza": round(rng.uniform(0.80, 0.98), 3),
                    })
            else:
                fuente = "conjunto simulado"
                for ambito, items in ARRANQUE_SIMULADO.items():
                    for titulo, obligatorio in items:
                        propuesta[ambito].append({
                            "titulo": titulo, "obligatorio": obligatorio,
                            "confianza": round(rng.uniform(0.78, 0.97), 3),
                        })

            total = sum(len(v) for v in propuesta.values())
            confianzas = [i["confianza"] for v in propuesta.values() for i in v]
            global_ = round(sum(confianzas) / len(confianzas), 3) if confianzas else 0.0

            review.campos_extraidos = {
                **base, **propuesta, "contrato_id": str(contrato_id),
                "confianza": global_, "fuente": fuente,
                "_accion_aplicada": (
                    f"{total} requisitos propuestos desde {fuente}; "
                    "confírmelos con POST /contratos/{id}/requisitos?bulk=true"),
            }
            review.confianza = global_
            review.resultado = "validado" if global_ >= 0.85 else "con_observaciones"
            review.status = "done"
            review.finished_at = datetime.now(timezone.utc)
            _reescribir_hallazgos(db, review, [
                {"tipo": "info", "codigo": "ARRANQUE_LEIDO",
                 "mensaje": (f"{total} requisitos detectados ({fuente}): "
                             f"{len(propuesta['empresa'])} de empresa, "
                             f"{len(propuesta['personal'])} de personal, "
                             f"{len(propuesta['equipo'])} de equipo.")},
            ] + ([] if titulos else [
                {"tipo": "warning", "codigo": "ARCHIVO_NO_TABULAR",
                 "mensaje": ("No se pudo leer una lista de requisitos del archivo; "
                             "se propone el conjunto base de carpeta de arranque.")},
            ]))
            db.commit()
        except Exception as e:  # noqa: BLE001
            logger.exception("extracción de carpeta de arranque %s falló", review_id)
            _fallar(db, rid, f"{type(e).__name__}: {e}")


# ============================================================================
# §15.3 — Generación de reportes
# ============================================================================
def _texto_pdf(s: str) -> bytes:
    """Escapa un literal de cadena PDF y lo lleva a latin-1 (fuente Type1)."""
    limpio = (str(s).replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)"))
    return limpio.encode("latin-1", "replace")


def pdf_simple(titulo: str, lineas: list[str]) -> bytes:
    """PDF A4 horizontal de texto monoespaciado, sin dependencias externas.

    La especificación prevé WeasyPrint o ReportLab (§15.3) y ninguno está en
    `requirements.txt`. Se intenta ReportLab por si el despliegue lo trae y, si
    no, se emite este PDF mínimo: válido y legible, pero sin tablas ni estilos.
    Se prefiere generar un PDF pobre a fallar el job, porque el formato lo eligió
    el usuario y el enum `reporte_formato` no admite degradar a CSV.
    """
    try:
        from reportlab.lib.pagesizes import landscape, A4      # import perezoso
        from reportlab.pdfgen import canvas
    except ImportError:
        return _pdf_minimo(titulo, lineas)

    buf = io.BytesIO()
    ancho, alto = landscape(A4)
    c = canvas.Canvas(buf, pagesize=(ancho, alto))
    y = alto - 40
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, titulo)
    y -= 20
    c.setFont("Courier", 8)
    for linea in lineas:
        if y < 40:
            c.showPage()
            c.setFont("Courier", 8)
            y = alto - 40
        c.drawString(40, y, linea[:150])
        y -= 11
    c.save()
    return buf.getvalue()


def _pdf_minimo(titulo: str, lineas: list[str], *, ancho: int = 842,
                alto: int = 595, por_pagina: int = 44) -> bytes:
    """Escritor PDF 1.4 de una sola fuente Courier. Suficiente para un listado."""
    cuerpo = [titulo, ""] + list(lineas)
    paginas = [cuerpo[i:i + por_pagina]
               for i in range(0, len(cuerpo), por_pagina)] or [[]]
    n = len(paginas)

    objetos: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        ("<< /Type /Pages /Kids [%s] /Count %d >>"
         % (" ".join(f"{4 + 2 * i} 0 R" for i in range(n)), n)).encode(),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>",
    ]
    for i, pagina in enumerate(paginas):
        objetos.append((
            "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 %d %d] "
            "/Resources << /Font << /F1 3 0 R >> >> /Contents %d 0 R >>"
            % (ancho, alto, 5 + 2 * i)).encode())
        partes = [b"BT /F1 9 Tf 11 TL 36 ", str(alto - 40).encode(), b" Td\n"]
        for linea in pagina:
            partes += [b"(", _texto_pdf(linea[:150]), b") Tj T*\n"]
        partes.append(b"ET")
        flujo = b"".join(partes)
        objetos.append(b"<< /Length %d >>\nstream\n%s\nendstream"
                       % (len(flujo), flujo))

    salida = bytearray(b"%PDF-1.4\n")
    offsets: list[int] = []
    for i, obj in enumerate(objetos, start=1):
        offsets.append(len(salida))
        salida += b"%d 0 obj\n%s\nendobj\n" % (i, obj)
    inicio_xref = len(salida)
    salida += b"xref\n0 %d\n0000000000 65535 f \n" % (len(objetos) + 1)
    for off in offsets:
        salida += b"%010d 00000 n \n" % off
    salida += (b"trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF\n"
               % (len(objetos) + 1, inicio_xref))
    return bytes(salida)


def _lineas_tabla(cabeceras: list[str], filas: list[list],
                  ancho_max: int = 148) -> list[str]:
    """Renderiza la tabla como texto de ancho fijo para el PDF."""
    cols = len(cabeceras)
    if not cols:
        return []
    anchos = [len(str(c)) for c in cabeceras]
    for fila in filas[:2000]:
        for i in range(cols):
            valor = "" if i >= len(fila) or fila[i] is None else str(fila[i])
            anchos[i] = max(anchos[i], min(len(valor), 28))
    # Reparto proporcional si la suma excede el ancho de página.
    total = sum(anchos) + 3 * (cols - 1)
    if total > ancho_max:
        factor = (ancho_max - 3 * (cols - 1)) / max(sum(anchos), 1)
        anchos = [max(6, int(a * factor)) for a in anchos]

    def linea(valores) -> str:
        celdas = []
        for i in range(cols):
            v = "" if i >= len(valores) or valores[i] is None else str(valores[i])
            celdas.append(v[:anchos[i]].ljust(anchos[i]))
        return " | ".join(celdas)

    salida = [linea(cabeceras), "-" * min(total, ancho_max)]
    salida += [linea(f) for f in filas]
    return salida


@tarea("generar_reporte")
def generar_reporte(reporte_id: str, company_id: str) -> None:
    """Materializa un `Reporte` encolado y lo deja disponible para descarga.

    Las filas se obtienen del mismo helper que usa `POST /exportaciones`
    (`routers.reportes.filas_de_recurso`, import perezoso para no crear un ciclo
    router↔worker), de modo que el reporte y la exportación de la misma vista no
    puedan divergir.

    El recurso efectivo sale de `params['recurso']` si está —lo escribe
    /exportaciones al derivar una exportación grande— y si no de `RECURSO_POR_TIPO`.
    """
    cid = uuid.UUID(company_id)
    with worker_session(company_id=cid) as db:
        rep = db.get(Reporte, uuid.UUID(reporte_id))
        if rep is None or rep.company_id != cid:
            logger.error("reporte %s inexistente o de otra empresa", reporte_id)
            return
        if rep.status == "done" and rep.blob_path:
            logger.info("reporte %s ya generado", rep.id)
            return
        rep.status = "processing"
        db.commit()

        try:
            from ..routers.reportes import filas_de_recurso, serializar

            params = dict(rep.params or {})
            recurso = params.get("recurso") or RECURSO_POR_TIPO.get(rep.tipo)
            if not recurso:
                raise ValueError(f"No hay recurso asociado al tipo '{rep.tipo}'")
            filtros = params.get("filtros")
            if not isinstance(filtros, dict):
                # Los params de /reportes son planos ({dias, contrato_id, ...}).
                filtros = {k: v for k, v in params.items()
                           if k not in ("recurso", "formato_export")}
            scope = filtros.get("contrato_id")
            scope_contrato = uuid.UUID(str(scope)) if scope else None

            cabeceras, filas = filas_de_recurso(db, cid, recurso, filtros,
                                                scope_contrato)

            formato = params.get("formato_export") or rep.formato
            if formato == "pdf":
                data = pdf_simple(rep.nombre, _lineas_tabla(cabeceras, filas))
                ext = "pdf"
            else:
                data, ext = serializar(formato, cabeceras, filas, rep.nombre)

            blob_path = f"reportes/{cid}/{rep.id}.{ext}"
            get_storage().save(blob_path, data)

            rep.blob_path = blob_path
            rep.status = "done"
            actividad.log(db, cid, "creacion", "reportes",
                          f"Reporte '{rep.nombre}' generado ({len(filas)} filas, "
                          f"{ext.upper()})", user_id=rep.generado_por,
                          entidad_tipo="reporte", entidad_id=rep.id)
            db.commit()
            logger.info("reporte %s generado en %s (%d filas)", rep.id,
                        blob_path, len(filas))
        except Exception as e:  # noqa: BLE001
            logger.exception("generación del reporte %s falló", reporte_id)
            db.rollback()
            rep = db.get(Reporte, uuid.UUID(reporte_id))
            if rep is not None:
                rep.status = "failed"
                params = dict(rep.params or {})
                params["error"] = f"{type(e).__name__}: {e}"[:500]
                rep.params = params
                db.commit()


# ============================================================================
# Notificaciones
# ============================================================================
@tarea("notificar_solicitud_acceso")
def notificar_solicitud_acceso(company_id: str,
                               contrato_plataforma_id: str | None = None,
                               **extra) -> None:
    """Cursa la solicitud de acceso a una plataforma del mandante.

    Hoy solo deja el rastro en `actividad`: no hay conector implementado y una
    solicitud perdida es peor que una registrada a mano, así que la tarea nunca
    falla por falta de integración.

    Acepta kwargs extra porque `routers/plataformas.py` encola con
    `plataforma_id`, `plataforma`, `integracion_id` y `nota`; ambos nombres del
    identificador de la plataforma se admiten para no romper ese llamador.
    """
    cid = uuid.UUID(company_id)
    pid = contrato_plataforma_id or extra.get("plataforma_id")
    nombre = extra.get("plataforma") or "plataforma"
    nota = (extra.get("nota") or "").strip()
    with worker_session(company_id=cid) as db:
        try:
            descripcion = f"Solicitud de acceso a '{nombre}' enviada al mandante"
            if nota:
                descripcion += f" — {nota}"
            actividad.log(
                db, cid, "actualizacion", "plataformas", descripcion,
                user_id=None, entidad_tipo="contrato_plataforma",
                entidad_id=uuid.UUID(str(pid)) if pid else None,
                plataforma=str(nombre))
            db.commit()
            # TODO(integraciones §16): cursar la solicitud por el conector
            # correspondiente (SIGA/WorkMate/MetaContratas) y notificar por
            # email/WhatsApp según `notificacion_preferencias`. Al implementarlo,
            # registrar el resultado en `sync_logs` con la integración recibida
            # en extra['integracion_id'].
            logger.info("solicitud de acceso registrada: empresa=%s plataforma=%s",
                        cid, nombre)
        except Exception:  # noqa: BLE001
            logger.exception("no se pudo registrar la solicitud de acceso")
            db.rollback()


# ============================================================================
# Mantenimiento
# ============================================================================
@tarea("purgar_temporales")
def purgar_temporales(horas: int = 24) -> int:
    """Elimina los blobs temporales de extracción IA con más de `horas` de vida.

    Los `tmp/{company_id}/...` los crea el flujo de extracción (cédula, padrón,
    contrato, carpeta de arranque) y no quedan referenciados por ninguna fila:
    sin purga crecen indefinidamente y contienen datos personales (cédulas), que
    es justo lo que la retención del §19 obliga a acotar.

    Con `STORAGE_BACKEND=azure` conviene además una regla de ciclo de vida en la
    cuenta de almacenamiento: es más barata y no depende de que el worker corra.
    """
    import time

    from ..config import settings
    from .storage import LocalStorage, get_storage

    st = get_storage()
    limite = time.time() - horas * 3600
    borrados = 0

    if isinstance(st, LocalStorage):
        base = os.path.join(st.base, "tmp")
        if not os.path.isdir(base):
            return 0
        for raiz, _dirs, archivos in os.walk(base):
            for nombre in archivos:
                ruta = os.path.join(raiz, nombre)
                try:
                    if os.path.getmtime(ruta) < limite:
                        os.remove(ruta)
                        borrados += 1
                except OSError:
                    continue
        return borrados

    try:
        contenedor = st.client.get_container_client(settings.azure_blob_container)
        for blob in contenedor.list_blobs(name_starts_with="tmp/"):
            creado = getattr(blob, "creation_time", None) or blob.last_modified
            if creado and creado.timestamp() < limite:
                st.delete(blob.name)
                borrados += 1
    except Exception:  # noqa: BLE001
        logger.exception("no se pudieron purgar los blobs temporales")
    return borrados
