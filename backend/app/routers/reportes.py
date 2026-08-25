"""Reportes, programaciones y exportaciones de vistas (§15.3).

Tres recursos con la misma materia prima y tres tiempos distintos:

* `POST /reportes` — job asíncrono. Devuelve 202 y el archivo se descarga luego
  con `GET /reportes/{id}/download-url`. Es el camino para PDF y para volúmenes
  grandes.
* `POST /reportes/programados` — el mismo job, disparado por cron. La ejecución
  la agenda el beat del worker; aquí solo se guarda la definición.
* `POST /exportaciones` — «bajar a Excel la tabla que estoy viendo». Hasta
  `EXPORT_FILAS_MAX` filas se genera en el propio request y se responde 200 con
  la URL de descarga; por encima se crea un `Reporte` equivalente y se responde
  202. El umbral existe porque una exportación bloquea un worker de la API: es
  aceptable para una tabla en pantalla, no para el histórico completo.

El punto de unión es `filas_de_recurso()`: **una sola** función construye
(cabeceras, filas) para cada vista exportable, y la usan tanto la exportación
síncrona como la tarea `generar_reporte` del worker. Si divergieran, el mismo
listado saldría distinto según su tamaño.

Dos limitaciones conscientes:

* `reporte_tipo` en la BD no tiene un valor por vista (son cinco tipos de
  informe, no nueve recursos), así que al derivar una exportación grande el
  recurso real viaja en `params['recurso']` y el `tipo` queda como etiqueta. La
  tarea prioriza `params['recurso']`.
* `reporte_formato` solo admite `pdf|excel`. Una exportación en CSV se persiste
  como `excel` con `params['formato_export']='csv'`.
"""
from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from datetime import date, datetime, timedelta, timezone

from fastapi import APIRouter, Depends, Query
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, joinedload

from ..config import EXPORT_RECURSOS, REPORTE_TIPOS, settings
from ..deps import (Page, aplicar_orden, contrato_scope, err, get_company_id,
                    get_current_user, get_db, paginacion, require_company,
                    sobre)
from ..models import (Alerta, Contrato, Documento, DocumentoArchivo, Faena,
                      Reporte, ReporteProgramado, RequisitoTemplate, Sujeto,
                      User)
from ..services import actividad
from ..services.checklist import normalizar, stats_sujeto
from ..services.jobs import enqueue
from ..services.storage import get_storage, make_tmp_path

logger = logging.getLogger("acredittia.reportes")

router = APIRouter(prefix="/reportes", tags=["reportes"])
router_export = APIRouter(prefix="/exportaciones", tags=["reportes"])

FORMATOS_REPORTE: tuple[str, ...] = ("pdf", "excel")
FORMATOS_EXPORT: tuple[str, ...] = ("excel", "csv")
ESTADOS_JOB: tuple[str, ...] = ("queued", "processing", "done", "failed")

# Valores del enum `reporte_tipo` en la BD (01_esquema.sql). REPORTE_TIPOS de
# config añade 'matriz_cumplimiento', que el enum NO tiene: se valida contra
# config (lo pide el contrato) y además contra esta tupla, para devolver un 400
# entendible en vez de un 22P02 del driver.
TIPOS_PERSISTIBLES: tuple[str, ...] = (
    "estado_acreditacion", "cumplimiento_requisitos", "personal_acreditado",
    "equipos_vehiculos", "vencimientos",
)

# Tipo de informe con el que se persiste una exportación derivada. Solo es una
# etiqueta: el recurso efectivo lo guarda `params['recurso']`.
TIPO_POR_RECURSO: dict[str, str] = {
    "personal": "personal_acreditado",
    "personas": "personal_acreditado",
    "equipos": "equipos_vehiculos",
    "flota": "equipos_vehiculos",
    "documentos": "vencimientos",
    "requisitos": "cumplimiento_requisitos",
    "alertas": "estado_acreditacion",
    "contratos": "estado_acreditacion",
    "matriz": "cumplimiento_requisitos",
}

ORDEN_REPORTES = {"created_at", "updated_at", "nombre", "tipo", "formato",
                  "status"}
ORDEN_PROGRAMADOS = {"created_at", "updated_at", "nombre", "tipo", "formato",
                     "activo", "ultimo_run_at"}

# Orden estable de las columnas de la matriz (§8.4).
ORDEN_AMBITO = {"empresa": 0, "personal": 1, "equipo": 2, "emsipor": 3}

ACTOR_PROGRAMADO = {"nombre": "Sistema (Programado)"}


# ============================================================================
# Entradas
# ============================================================================
class ReporteIn(BaseModel):
    tipo: str
    formato: str
    params: dict = Field(default_factory=dict)
    nombre: str | None = None


class ProgramadoIn(BaseModel):
    nombre: str
    tipo: str
    formato: str
    params: dict = Field(default_factory=dict)
    cron_expr: str
    activo: bool = True


class ProgramadoPatch(BaseModel):
    nombre: str | None = None
    tipo: str | None = None
    formato: str | None = None
    params: dict | None = None
    cron_expr: str | None = None
    activo: bool | None = None


class ExportacionIn(BaseModel):
    recurso: str
    filtros: dict = Field(default_factory=dict)
    formato: str = "excel"


# ============================================================================
# Serialización de archivos
# ============================================================================
def csv_bytes(cabeceras: list[str], filas: list[list]) -> bytes:
    """CSV con `;` y BOM UTF-8.

    Excel en configuración regional chilena interpreta la coma como separador
    decimal y no como separador de campos; con `;` y BOM el archivo se abre de
    doble clic sin pasar por el asistente de importación.
    """
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL,
                   lineterminator="\r\n")
    w.writerow(cabeceras)
    for fila in filas:
        w.writerow(["" if v is None else v for v in fila])
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def excel_bytes(cabeceras: list[str], filas: list[list],
                titulo: str) -> bytes | None:
    """XLSX con openpyxl, o None si la librería no está instalada.

    El import es perezoso a propósito: `openpyxl` no está en requirements.txt y
    no se quiere convertir en dependencia dura de arrancar la API. Quien llama
    decide qué hacer con el None (aquí, degradar a CSV).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl no instalado; la exportación degrada a CSV")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = (re.sub(r"[\[\]:*?/\\]", "-", titulo) or "Reporte")[:31]
    ws.append(list(cabeceras))
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E79")
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    for fila in filas:
        ws.append(["" if v is None else v for v in fila])
    ws.freeze_panes = "A2"
    if cabeceras:
        ws.auto_filter.ref = (f"A1:{get_column_letter(len(cabeceras))}"
                              f"{max(ws.max_row, 1)}")
    # Ancho aproximado por el contenido de las primeras filas: recorrer 50.000
    # celdas para ajustar columnas no vale el tiempo de un request.
    for i, cab in enumerate(cabeceras, start=1):
        ancho = len(str(cab))
        for fila in filas[:200]:
            if i <= len(fila) and fila[i - 1] is not None:
                ancho = max(ancho, len(str(fila[i - 1])))
        ws.column_dimensions[get_column_letter(i)].width = min(max(ancho + 2, 10), 46)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def serializar(formato: str, cabeceras: list[str], filas: list[list],
               titulo: str) -> tuple[bytes, str]:
    """Devuelve (contenido, extensión). Si se pidió Excel y no hay openpyxl,
    degrada a CSV; quien llama detecta la degradación porque la extensión no
    coincide con el formato pedido."""
    if formato == "excel":
        data = excel_bytes(cabeceras, filas, titulo)
        if data is not None:
            return data, "xlsx"
        return csv_bytes(cabeceras, filas), "csv"
    return csv_bytes(cabeceras, filas), "csv"


def _slug(texto: str) -> str:
    limpio = re.sub(r"[^A-Za-z0-9]+", "-", texto or "export").strip("-").lower()
    return (limpio or "export")[:60]


def nombre_archivo(titulo: str, formato: str, ext: str) -> str:
    """Nombre visible de la descarga; avisa si el Excel degradó a CSV."""
    sufijo = "-excel-no-disponible" if (formato == "excel" and ext == "csv") else ""
    return f"{_slug(titulo)}-{date.today().isoformat()}{sufijo}.{ext}"


# ============================================================================
# Lectura de filtros (llegan como JSON libre: todo puede venir como string)
# ============================================================================
def _f_str(filtros: dict, clave: str) -> str | None:
    v = (filtros or {}).get(clave)
    if v is None or v == "":
        return None
    return str(v)


def _f_uuid(filtros: dict, clave: str) -> uuid.UUID | None:
    v = _f_str(filtros, clave)
    if v is None:
        return None
    try:
        return uuid.UUID(v)
    except ValueError:
        raise err(400, "FILTRO_INVALIDO", f"'{clave}' no es un UUID válido")


def _f_bool(filtros: dict, clave: str) -> bool | None:
    v = (filtros or {}).get(clave)
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("1", "true", "si", "sí", "yes")


def _f_int(filtros: dict, clave: str) -> int | None:
    v = _f_str(filtros, clave)
    if v is None:
        return None
    try:
        return int(float(v))
    except ValueError:
        raise err(400, "FILTRO_INVALIDO", f"'{clave}' debe ser numérico")


def _f_date(filtros: dict, clave: str) -> date | None:
    v = _f_str(filtros, clave)
    if v is None:
        return None
    try:
        return date.fromisoformat(v[:10])
    except ValueError:
        raise err(400, "FILTRO_INVALIDO", f"'{clave}' debe ser una fecha ISO")


def _si_no(v) -> str:
    return "Sí" if v else "No"


def _fecha(v: date | None) -> str:
    return v.isoformat() if v else ""


def _fecha_hora(v: datetime | None) -> str:
    return v.isoformat() if v else ""


# ============================================================================
# Consultas base por recurso
# ============================================================================
def _q_sujetos(cid: uuid.UUID, tipo: str, filtros: dict,
               scope: uuid.UUID | None):
    """Sujetos de un tipo con los filtros de la vista de personal / equipos."""
    q = (select(Sujeto)
         .join(Contrato, Contrato.id == Sujeto.contrato_id)
         .where(Sujeto.company_id == cid, Sujeto.tipo == tipo))
    contrato_id = _f_uuid(filtros, "contrato_id") or scope
    if contrato_id:
        q = q.where(Sujeto.contrato_id == contrato_id)
    faena_id = _f_uuid(filtros, "faena_id")
    if faena_id:
        q = q.where(Contrato.faena_id == faena_id)
    cargo_id = _f_uuid(filtros, "cargo_id")
    if cargo_id:
        q = q.where(Sujeto.cargo_id == cargo_id)
    estado = _f_str(filtros, "estado")
    if estado:
        q = q.where(Sujeto.estado == estado)
    else:
        # Una exportación de roster no debería arrastrar bajas salvo que se pidan.
        if not _f_bool(filtros, "incluir_bajas"):
            q = q.where(Sujeto.estado != "baja")
    conductor = _f_bool(filtros, "es_conductor")
    if conductor is not None:
        q = q.where(Sujeto.es_conductor.is_(conductor))
    tipo_equipo = _f_str(filtros, "tipo_equipo")
    if tipo_equipo:
        q = q.where(Sujeto.tipo_equipo == tipo_equipo)
    search = _f_str(filtros, "search")
    if search:
        like = f"%{search}%"
        q = q.where(or_(Sujeto.nombre.ilike(like), Sujeto.rut.ilike(like),
                        Sujeto.patente.ilike(like)))
    return q


def _q_documentos(cid: uuid.UUID, filtros: dict, scope: uuid.UUID | None):
    """Documentos de la empresa con los filtros de la vista de documentos.

    El dueño es `sujeto_id` XOR `contrato_id`, así que el filtro por contrato
    tiene que cubrir las dos rutas con un OR sobre el LEFT JOIN al sujeto.
    """
    q = (select(Documento)
         .outerjoin(Sujeto, Sujeto.id == Documento.sujeto_id)
         .where(Documento.company_id == cid))
    contrato_id = _f_uuid(filtros, "contrato_id") or scope
    if contrato_id:
        q = q.where(or_(Documento.contrato_id == contrato_id,
                        Sujeto.contrato_id == contrato_id))
    sujeto_id = _f_uuid(filtros, "sujeto_id")
    if sujeto_id:
        q = q.where(Documento.sujeto_id == sujeto_id)
    estado = _f_str(filtros, "estado")
    if estado:
        q = q.where(Documento.estado == estado)
    estado_calc = _f_str(filtros, "estado_calc")
    if estado_calc:
        q = q.where(Documento.estado_calc == estado_calc)
    obligatorio = _f_bool(filtros, "obligatorio")
    if obligatorio is not None:
        q = q.where(Documento.obligatorio.is_(obligatorio))
    emsipor = _f_bool(filtros, "es_emsipor")
    if emsipor is not None:
        q = q.where(Documento.es_emsipor.is_(emsipor))
    dias = _f_int(filtros, "dias")
    if dias is not None:
        # Ventana de vencimientos: es el filtro del reporte 'vencimientos'.
        q = q.where(Documento.vence.is_not(None),
                    Documento.vence <= date.today() + timedelta(days=dias))
    desde, hasta = _f_date(filtros, "desde"), _f_date(filtros, "hasta")
    if desde:
        q = q.where(Documento.vence >= desde)
    if hasta:
        q = q.where(Documento.vence <= hasta)
    search = _f_str(filtros, "search")
    if search:
        q = q.where(Documento.titulo.ilike(f"%{search}%"))
    return q


def _q_alertas(cid: uuid.UUID, filtros: dict, scope: uuid.UUID | None):
    q = select(Alerta).where(Alerta.company_id == cid)
    contrato_id = _f_uuid(filtros, "contrato_id") or scope
    if contrato_id:
        q = q.where(Alerta.contrato_id == contrato_id)
    for campo, columna in (("severidad", Alerta.severidad),
                           ("estado", Alerta.estado),
                           ("origen", Alerta.origen)):
        valor = _f_str(filtros, campo)
        if valor:
            q = q.where(columna == valor)
    if _f_bool(filtros, "solo_activas"):
        q = q.where(Alerta.resuelta_at.is_(None))
    leida = _f_bool(filtros, "leida")
    if leida is not None:
        q = q.where(Alerta.leida_at.is_not(None) if leida
                    else Alerta.leida_at.is_(None))
    desde, hasta = _f_date(filtros, "desde"), _f_date(filtros, "hasta")
    if desde:
        q = q.where(Alerta.created_at >= datetime.combine(
            desde, datetime.min.time(), tzinfo=timezone.utc))
    if hasta:
        q = q.where(Alerta.created_at < datetime.combine(
            hasta + timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc))
    search = _f_str(filtros, "search")
    if search:
        q = q.where(Alerta.titulo.ilike(f"%{search}%"))
    return q


def _q_contratos(cid: uuid.UUID, filtros: dict, scope: uuid.UUID | None):
    q = (select(Contrato).join(Faena, Faena.id == Contrato.faena_id)
         .where(Contrato.company_id == cid))
    contrato_id = _f_uuid(filtros, "contrato_id") or scope
    if contrato_id:
        q = q.where(Contrato.id == contrato_id)
    faena_id = _f_uuid(filtros, "faena_id")
    if faena_id:
        q = q.where(Contrato.faena_id == faena_id)
    estado = _f_str(filtros, "estado")
    if estado:
        q = q.where(Contrato.estado == estado)
    search = _f_str(filtros, "search")
    if search:
        like = f"%{search}%"
        q = q.where(or_(Contrato.nombre.ilike(like), Contrato.codigo.ilike(like),
                        Faena.nombre.ilike(like)))
    return q


def _q_templates(cid: uuid.UUID, filtros: dict, scope: uuid.UUID | None):
    """Catálogo de requisitos visible para la empresa (mismo criterio que §15.1)."""
    from .requisitos import _catalogo_visible          # se reutiliza la visibilidad
    q = _catalogo_visible(cid, scope)
    ambito = _f_str(filtros, "ambito")
    if ambito:
        q = q.where(RequisitoTemplate.ambito == ambito)
    tipo = _f_str(filtros, "tipo")
    if tipo:
        q = q.where(RequisitoTemplate.tipo == tipo)
    obligatorio = _f_bool(filtros, "obligatorio")
    if obligatorio is not None:
        q = q.where(RequisitoTemplate.obligatorio.is_(obligatorio))
    faena_id = _f_uuid(filtros, "faena_id")
    if faena_id:
        q = q.where(RequisitoTemplate.faena_id == faena_id)
    search = _f_str(filtros, "search")
    if search:
        q = q.where(RequisitoTemplate.titulo.ilike(f"%{search}%"))
    return q


# ============================================================================
# Datos derivados compartidos
# ============================================================================
def _docs_por_sujeto(db: Session, cid: uuid.UUID,
                     ids: list[uuid.UUID]) -> dict[uuid.UUID, list[Documento]]:
    if not ids:
        return {}
    filas = db.scalars(select(Documento).where(
        Documento.company_id == cid,
        Documento.sujeto_id.in_(ids)).order_by(Documento.titulo)).all()
    salida: dict[uuid.UUID, list[Documento]] = {}
    for d in filas:
        salida.setdefault(d.sujeto_id, []).append(d)
    return salida


def _proximo(docs: list[Documento]) -> str:
    """Fecha del próximo vencimiento aún futuro, o el vencido más antiguo."""
    fechas = sorted(d.vence for d in docs if d.vence)
    if not fechas:
        return ""
    hoy = date.today()
    futuras = [f for f in fechas if f >= hoy]
    return (futuras[0] if futuras else fechas[0]).isoformat()


def _contratos_por_id(db: Session, cid: uuid.UUID) -> dict[uuid.UUID, tuple]:
    """(nombre, faena, mandante) por contrato, para no consultar fila a fila."""
    filas = db.execute(
        select(Contrato.id, Contrato.nombre, Faena.nombre, Faena.mandante)
        .join(Faena, Faena.id == Contrato.faena_id)
        .where(Contrato.company_id == cid)).all()
    return {f[0]: (f[1], f[2], f[3]) for f in filas}


def _archivos_por_documento(db: Session, cid: uuid.UUID,
                            ids: list[uuid.UUID]) -> dict[uuid.UUID, int]:
    if not ids:
        return {}
    filas = db.execute(
        select(DocumentoArchivo.documento_id, func.count())
        .where(DocumentoArchivo.company_id == cid,
               DocumentoArchivo.documento_id.in_(ids))
        .group_by(DocumentoArchivo.documento_id)).all()
    return {f[0]: int(f[1]) for f in filas}


def _ambito_doc(d: Documento, sujeto: Sujeto | None) -> str:
    if d.es_emsipor:
        return "emsipor"
    if d.contrato_id:
        return "empresa"
    if sujeto is not None:
        return "equipo" if sujeto.tipo == "equipo" else "personal"
    return "personal"


# ============================================================================
# Constructores de filas por recurso
# ============================================================================
def _filas_personal(db: Session, cid: uuid.UUID, filtros: dict,
                    scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    q = _q_sujetos(cid, "trabajador", filtros, scope).order_by(Sujeto.nombre)
    sujetos = list(db.scalars(q.options(
        joinedload(Sujeto.contrato).joinedload(Contrato.faena))).unique())
    docs = _docs_por_sujeto(db, cid, [s.id for s in sujetos])
    cab = ["Nombre", "RUT", "Cargo", "Conductor", "Estado", "Contrato", "Faena",
           "Mandante", "Docs OK", "Docs total", "Vencidos", "Por vencer",
           "Cumplimiento %", "Próximo vencimiento"]
    filas = []
    for s in sujetos:
        d = docs.get(s.id, [])
        st = stats_sujeto(d)
        c, f = s.contrato, (s.contrato.faena if s.contrato else None)
        filas.append([
            s.nombre, s.rut or "", s.cargo or "", _si_no(s.es_conductor),
            s.estado, c.nombre if c else "", f.nombre if f else "",
            f.mandante if f else "", st["ok"], st["docs_total"], st["vencidos"],
            st["por_vencer"], st["cumplimiento_pct"], _proximo(d),
        ])
    return cab, filas


def _filas_equipos(db: Session, cid: uuid.UUID, filtros: dict,
                   scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    q = _q_sujetos(cid, "equipo", filtros, scope).order_by(Sujeto.patente)
    sujetos = list(db.scalars(q.options(
        joinedload(Sujeto.contrato).joinedload(Contrato.faena))).unique())
    docs = _docs_por_sujeto(db, cid, [s.id for s in sujetos])
    cab = ["Patente", "Tipo de equipo", "Marca", "Modelo", "Año", "Estado",
           "Contrato", "Faena", "Mandante", "Docs OK", "Docs total",
           "Vencidos", "Por vencer", "Cumplimiento %", "Próximo vencimiento"]
    filas = []
    for s in sujetos:
        d = docs.get(s.id, [])
        st = stats_sujeto(d)
        c, f = s.contrato, (s.contrato.faena if s.contrato else None)
        filas.append([
            s.patente or "", s.tipo_equipo or "", s.marca or "", s.modelo or "",
            s.anio or "", s.estado, c.nombre if c else "",
            f.nombre if f else "", f.mandante if f else "", st["ok"],
            st["docs_total"], st["vencidos"], st["por_vencer"],
            st["cumplimiento_pct"], _proximo(d),
        ])
    return cab, filas


def _agrupar_identidad(db: Session, cid: uuid.UUID, tipo: str, filtros: dict,
                       scope: uuid.UUID | None) -> list[tuple[str, list[Sujeto]]]:
    """Registros agrupados por RUT o patente, en orden alfabético de la clave.

    Una persona no es un `Sujeto`: es la identidad que agrupa sus registros en
    varios contratos (§9.1). La exportación de `/personas` y `/flota` refleja esa
    agrupación, no la tabla.
    """
    campo = Sujeto.rut if tipo == "trabajador" else Sujeto.patente
    q = _q_sujetos(cid, tipo, filtros, scope).where(campo.is_not(None), campo != "")
    sujetos = list(db.scalars(q.options(
        joinedload(Sujeto.contrato).joinedload(Contrato.faena))).unique())
    grupos: dict[str, list[Sujeto]] = {}
    for s in sujetos:
        grupos.setdefault(s.rut if tipo == "trabajador" else s.patente, []).append(s)
    return sorted(grupos.items(), key=lambda kv: kv[0] or "")


def _union_stats(docs: list[Documento]) -> dict:
    """Cumplimiento sobre la UNIÓN de documentos obligatorios, por título.

    El mismo requisito existe repetido en cada contrato; para la identidad basta
    tenerlo vigente en uno para considerarlo cubierto (misma regla que §9.1).
    """
    mejor: dict[str, Documento] = {}
    peso = {"ok": 0, "porvenc": 1, "venc": 2, "falta": 3}
    for d in docs:
        if d.es_emsipor:
            continue
        clave = d.titulo.strip().lower()
        actual = mejor.get(clave)
        if actual is None or peso.get(d.estado_calc, 9) < peso.get(actual.estado_calc, 9):
            mejor[clave] = d
    return stats_sujeto(list(mejor.values()))


def _filas_personas(db: Session, cid: uuid.UUID, filtros: dict,
                    scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    grupos = _agrupar_identidad(db, cid, "trabajador", filtros, scope)
    docs = _docs_por_sujeto(db, cid, [s.id for _, g in grupos for s in g])
    cab = ["RUT", "Nombre", "Registros", "Contratos", "Cargos", "Estados",
           "Docs OK (unión)", "Docs total (unión)", "Cumplimiento %",
           "Próximo vencimiento"]
    filas = []
    for clave, g in grupos:
        todos = [d for s in g for d in docs.get(s.id, [])]
        st = _union_stats(todos)
        filas.append([
            clave, g[0].nombre, len(g),
            "; ".join(sorted({s.contrato.nombre for s in g if s.contrato})),
            "; ".join(sorted({s.cargo for s in g if s.cargo})),
            "; ".join(sorted({s.estado for s in g})),
            st["ok"], st["docs_total"], st["cumplimiento_pct"], _proximo(todos),
        ])
    return cab, filas


def _filas_flota(db: Session, cid: uuid.UUID, filtros: dict,
                 scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    grupos = _agrupar_identidad(db, cid, "equipo", filtros, scope)
    docs = _docs_por_sujeto(db, cid, [s.id for _, g in grupos for s in g])
    cab = ["Patente", "Tipo de equipo", "Marca", "Modelo", "Año", "Registros",
           "Contratos", "Estados", "Docs OK (unión)", "Docs total (unión)",
           "Cumplimiento %", "Próximo vencimiento"]
    filas = []
    for clave, g in grupos:
        todos = [d for s in g for d in docs.get(s.id, [])]
        st = _union_stats(todos)
        ref = g[0]
        filas.append([
            clave, ref.tipo_equipo or "", ref.marca or "", ref.modelo or "",
            ref.anio or "", len(g),
            "; ".join(sorted({s.contrato.nombre for s in g if s.contrato})),
            "; ".join(sorted({s.estado for s in g})),
            st["ok"], st["docs_total"], st["cumplimiento_pct"], _proximo(todos),
        ])
    return cab, filas


def _filas_documentos(db: Session, cid: uuid.UUID, filtros: dict,
                      scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    q = _q_documentos(cid, filtros, scope).order_by(
        Documento.vence.is_(None), Documento.vence, Documento.titulo)
    docs = list(db.scalars(q.options(joinedload(Documento.sujeto))).unique())
    contratos = _contratos_por_id(db, cid)
    archivos = _archivos_por_documento(db, cid, [d.id for d in docs])
    hoy = date.today()
    cab = ["Documento", "Ámbito", "Dueño", "RUT / Patente", "Contrato", "Faena",
           "Mandante", "Obligatorio", "EMSIPOR", "Estado", "Estado calculado",
           "Vence", "Días para vencer", "Archivos"]
    filas = []
    for d in docs:
        s = d.sujeto
        contrato_id = d.contrato_id or (s.contrato_id if s else None)
        nombre_c, faena, mandante = contratos.get(contrato_id, ("", "", ""))
        filas.append([
            d.titulo, _ambito_doc(d, s),
            s.nombre if s else (nombre_c or "Empresa"),
            (s.rut or s.patente or "") if s else "",
            nombre_c, faena, mandante,
            _si_no(d.obligatorio), _si_no(d.es_emsipor), d.estado,
            d.estado_calc, _fecha(d.vence),
            (d.vence - hoy).days if d.vence else "",
            archivos.get(d.id, 0),
        ])
    return cab, filas


def _filas_requisitos(db: Session, cid: uuid.UUID, filtros: dict,
                      scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    from .requisitos import _conteos_por_template, _estado_agregado

    q = _q_templates(cid, filtros, scope).order_by(
        RequisitoTemplate.ambito, RequisitoTemplate.titulo)
    templates = list(db.scalars(q))
    conteos = _conteos_por_template(db, cid, [t.id for t in templates], scope)
    faenas = {f.id: f.nombre for f in db.scalars(select(Faena))}
    cab = ["Requisito", "Código", "Ámbito", "Tipo", "Obligatorio",
           "Vigencia (meses)", "Plataforma", "Faena", "Documentos", "OK",
           "Por vencer", "Vencidos", "Falta", "Estado agregado",
           "Cumplimiento %"]
    filas = []
    for t in templates:
        c = conteos.get(t.id, {"docs": 0, "ok": 0, "porvenc": 0, "venc": 0,
                               "falta": 0})
        pct = round(100 * c["ok"] / c["docs"]) if c["docs"] else 0
        filas.append([
            t.titulo, t.codigo or "", t.ambito, t.tipo or "",
            _si_no(t.obligatorio), t.vigencia_meses or "", t.plataforma or "",
            faenas.get(t.faena_id, "Todas"), c["docs"], c["ok"], c["porvenc"],
            c["venc"], c["falta"], _estado_agregado(c), pct,
        ])
    return cab, filas


def _filas_alertas(db: Session, cid: uuid.UUID, filtros: dict,
                   scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    q = _q_alertas(cid, filtros, scope).order_by(Alerta.created_at.desc())
    alertas = list(db.scalars(q))
    contratos = _contratos_por_id(db, cid)
    ids = [a.sujeto_id for a in alertas if a.sujeto_id]
    sujetos = {s.id: s for s in db.scalars(select(Sujeto).where(
        Sujeto.company_id == cid, Sujeto.id.in_(ids)))} if ids else {}
    cab = ["Severidad", "Estado", "Origen", "Título", "Descripción",
           "Plataforma", "Contrato", "Faena", "Sujeto", "Leída", "Resuelta",
           "Creada"]
    filas = []
    for a in alertas:
        nombre_c, faena, _ = contratos.get(a.contrato_id, ("", "", ""))
        s = sujetos.get(a.sujeto_id)
        filas.append([
            a.severidad, a.estado, a.origen, a.titulo, a.descripcion or "",
            a.plataforma or "", nombre_c, faena,
            (s.nombre or s.patente or "") if s else "",
            _si_no(a.leida_at), _si_no(a.resuelta_at),
            _fecha_hora(a.created_at),
        ])
    return cab, filas


def _filas_contratos(db: Session, cid: uuid.UUID, filtros: dict,
                     scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    from .contratos import _stats                      # se reutiliza el cálculo

    q = _q_contratos(cid, filtros, scope).order_by(Contrato.nombre)
    contratos = list(db.scalars(q.options(joinedload(Contrato.faena))).unique())
    cab = ["Contrato", "Código", "Faena", "Mandante", "Grupo", "Estado",
           "Inicio", "Término", "Renovación automática", "Personal acreditado",
           "Personal total", "Equipos acreditados", "Equipos total",
           "Docs empresa OK", "Docs empresa total", "Cumplimiento %",
           "Alertas activas"]
    filas = []
    for c in contratos:
        st = _stats(db, c)
        f = c.faena
        filas.append([
            c.nombre, c.codigo or "", f.nombre if f else "",
            f.mandante if f else "", (f.grupo if f else "") or "", c.estado,
            _fecha(c.fecha_inicio), _fecha(c.fecha_termino),
            _si_no(c.renovacion_automatica),
            st["personal"]["acreditados"], st["personal"]["total"],
            st["equipos"]["acreditados"], st["equipos"]["total"],
            st["docs_empresa"]["ok"], st["docs_empresa"]["total"],
            st["cumplimiento_pct"], st["alertas_activas"],
        ])
    return cab, filas


def _filas_matriz(db: Session, cid: uuid.UUID, filtros: dict,
                  scope: uuid.UUID | None) -> tuple[list[str], list[list]]:
    """Matriz sujeto × requisito (§8.4) aplanada a una tabla.

    La matriz es dispersa: un requisito que no aplica al sujeto se exporta como
    «—», igual que lo renderiza la cuadrícula. El orden de columnas es el mismo
    que en la API (ámbito, obligatorio descendente, título) para que la columna
    N del Excel sea la columna N de la pantalla.
    """
    contrato_id = _f_uuid(filtros, "contrato_id") or scope
    if not contrato_id:
        raise err(400, "CONTRATO_REQUERIDO",
                  "La matriz exige 'contrato_id' en los filtros")
    tipo_vista = (_f_str(filtros, "tipo") or "personal").lower()
    if tipo_vista not in ("personal", "equipo"):
        raise err(400, "TIPO_INVALIDO", "tipo debe ser 'personal' o 'equipo'")
    incluir_opcionales = _f_bool(filtros, "incluir_opcionales")

    sub = dict(filtros or {})
    sub["contrato_id"] = str(contrato_id)
    q = _q_sujetos(cid, "trabajador" if tipo_vista == "personal" else "equipo",
                   sub, scope)
    sujetos = list(db.scalars(q.order_by(
        Sujeto.nombre if tipo_vista == "personal" else Sujeto.patente)))
    docs = _docs_por_sujeto(db, cid, [s.id for s in sujetos])

    columnas: dict[str, tuple] = {}
    for s in sujetos:
        for d in docs.get(s.id, []):
            if not d.obligatorio and not incluir_opcionales:
                continue
            clave = d.titulo.strip().lower()
            if clave not in columnas:
                # Misma clave de orden que `contratos.matriz`: el título
                # normalizado (sin tildes ni mayúsculas) para que la secuencia
                # sea alfabética en español y para que la columna N del Excel
                # sea la columna N de la pantalla.
                columnas[clave] = (ORDEN_AMBITO.get(_ambito_doc(d, s), 9),
                                   0 if d.obligatorio else 1,
                                   normalizar(d.titulo), d.titulo)
    orden = sorted(columnas.items(), key=lambda kv: kv[1][:3])
    titulos = [v[3] for _, v in orden]
    claves = [k for k, _ in orden]

    etiqueta = "Cargo" if tipo_vista == "personal" else "Tipo de equipo"
    cab = ["Sujeto", "RUT / Patente", etiqueta, "Cumplimiento %"] + titulos
    filas = []
    for s in sujetos:
        propios = {d.titulo.strip().lower(): d for d in docs.get(s.id, [])}
        st = stats_sujeto(docs.get(s.id, []))
        fila = [s.nombre if tipo_vista == "personal" else (s.patente or ""),
                (s.rut or s.patente or ""),
                (s.cargo or "") if tipo_vista == "personal" else (s.tipo_equipo or ""),
                st["cumplimiento_pct"]]
        for k in claves:
            d = propios.get(k)
            if d is None:
                fila.append("—")           # requisito no aplicable al sujeto
            elif d.vence:
                fila.append(f"{d.estado_calc} ({d.vence.isoformat()})")
            else:
                fila.append(d.estado_calc)
        filas.append(fila)
    return cab, filas


CONSTRUCTORES = {
    "personal": _filas_personal,
    "equipos": _filas_equipos,
    "personas": _filas_personas,
    "flota": _filas_flota,
    "documentos": _filas_documentos,
    "requisitos": _filas_requisitos,
    "alertas": _filas_alertas,
    "contratos": _filas_contratos,
    "matriz": _filas_matriz,
}


def filas_de_recurso(db: Session, cid: uuid.UUID, recurso: str, filtros: dict,
                     scope_contrato: uuid.UUID | None
                     ) -> tuple[list[str], list[list]]:
    """(cabeceras, filas) de una vista exportable.

    Único punto de verdad de las exportaciones: lo usa `POST /exportaciones`
    para generar el archivo en el request y la tarea `generar_reporte` del worker
    para generarlo en diferido. `scope_contrato` acota la salida a un contrato
    (el del contract_admin, o el que venga en los filtros del reporte).
    """
    fn = CONSTRUCTORES.get(recurso)
    if fn is None:
        raise err(400, "RECURSO_INVALIDO",
                  f"Recurso debe ser uno de: {', '.join(sorted(CONSTRUCTORES))}")
    return fn(db, cid, filtros or {}, scope_contrato)


def contar_de_recurso(db: Session, cid: uuid.UUID, recurso: str, filtros: dict,
                      scope_contrato: uuid.UUID | None) -> int:
    """Filas que produciría `filas_de_recurso`, sin materializarlas.

    Los recursos que se pueden contar en SQL lo hacen con un COUNT; los que
    agregan en memoria (identidades, catálogo, matriz) se cuentan construyendo,
    porque el conteo exacto requiere el mismo trabajo. Se separa igualmente para
    que la ruta rápida —la mayoría de las exportaciones— no cargue objetos ORM
    solo para descubrir que hay que derivar a un job.
    """
    filtros = filtros or {}
    consultas = {
        "personal": lambda: _q_sujetos(cid, "trabajador", filtros, scope_contrato),
        "equipos": lambda: _q_sujetos(cid, "equipo", filtros, scope_contrato),
        "documentos": lambda: _q_documentos(cid, filtros, scope_contrato),
        "alertas": lambda: _q_alertas(cid, filtros, scope_contrato),
        "contratos": lambda: _q_contratos(cid, filtros, scope_contrato),
    }
    if recurso in consultas:
        sub = consultas[recurso]().subquery()
        return db.scalar(select(func.count()).select_from(sub)) or 0
    _, filas = filas_de_recurso(db, cid, recurso, filtros, scope_contrato)
    return len(filas)


# ============================================================================
# Validaciones
# ============================================================================
def _solo_empresa(user: User) -> None:
    """Reportes y exportaciones son de alcance empresa (§7.1).

    El administrador de contrato no accede a este módulo: su vista es un contrato
    y un reporte agregado de la empresa le mostraría datos de los demás.
    """
    if user.role == "contract_admin":
        raise err(403, "ROL_INSUFICIENTE",
                  "El administrador de contrato no accede a reportes")


def _valida_tipo(tipo: str) -> str:
    if tipo not in REPORTE_TIPOS:
        raise err(400, "TIPO_INVALIDO",
                  f"Tipo debe ser uno de: {', '.join(REPORTE_TIPOS)}")
    if tipo not in TIPOS_PERSISTIBLES:
        raise err(400, "TIPO_NO_SOPORTADO",
                  f"El enum reporte_tipo de la base no admite '{tipo}' todavía; "
                  "expórtelo con POST /exportaciones (recurso='matriz')")
    return tipo


def _valida_formato(formato: str, permitidos: tuple[str, ...]) -> str:
    if formato not in permitidos:
        raise err(400, "FORMATO_INVALIDO",
                  f"Formato debe ser uno de: {', '.join(permitidos)}")
    return formato


def _campo_cron_valido(campo: str, lo: int, hi: int) -> bool:
    """Valida un campo cron con listas, rangos y pasos: `*`, `a`, `a-b`, `*/n`."""
    for parte in campo.split(","):
        if not parte:
            return False
        cuerpo, sep_paso, paso = parte.partition("/")
        if sep_paso:
            if not paso.isdigit() or not 1 <= int(paso) <= hi + 1:
                return False
        if cuerpo == "*":
            continue
        ini, sep, fin = cuerpo.partition("-")
        try:
            a = int(ini)
            b = int(fin) if sep else a
        except ValueError:
            return False
        if not (lo <= a <= hi and lo <= b <= hi and a <= b):
            return False
    return True


def valida_cron(expr: str) -> str:
    """Validación estructural de un cron de 5 campos (min hora dom mes dow).

    LIMITACIÓN CONOCIDA: no es un parser de cron. Comprueba que haya cinco
    campos separados por espacios y que cada uno use `*`, números, rangos,
    listas y pasos dentro de rangos plausibles. NO acepta nombres (`MON`,
    `JAN`), ni extensiones de Quartz (`L`, `W`, `#`, `?`), ni detecta fechas
    imposibles como `31 2 *` (31 de febrero). El parser real es el del beat del
    worker; esta validación solo evita guardar basura evidente.
    """
    campos = (expr or "").split()
    if len(campos) != 5:
        raise err(400, "CRON_INVALIDO",
                  "cron_expr debe tener 5 campos: minuto hora día-del-mes mes "
                  "día-de-semana (ej. '0 7 * * 1')")
    rangos = ((0, 59), (0, 23), (1, 31), (1, 12), (0, 7))
    nombres = ("minuto", "hora", "día del mes", "mes", "día de semana")
    for campo, (lo, hi), nombre in zip(campos, rangos, nombres):
        if not _campo_cron_valido(campo, lo, hi):
            raise err(400, "CRON_INVALIDO",
                      f"El campo '{nombre}' ('{campo}') no es válido; "
                      f"se esperaba *, un número {lo}-{hi}, un rango o un paso")
    return " ".join(campos)


# ============================================================================
# Serializadores de salida
# ============================================================================
def _generador_out(uid, nombre, email) -> dict:
    """Autor del reporte. Sin `generado_por` lo disparó una programación."""
    if uid is None:
        return dict(ACTOR_PROGRAMADO)
    return {"id": str(uid), "nombre": nombre, "email": email}


def _out(r: Reporte, nombre=None, email=None) -> dict:
    return {
        "id": str(r.id), "nombre": r.nombre, "tipo": r.tipo,
        "formato": r.formato, "status": r.status, "params": r.params or {},
        "descargable": bool(r.status == "done" and r.blob_path),
        "generado_por": _generador_out(r.generado_por, nombre, email),
        "created_at": r.created_at.isoformat(),
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
    }


def _out_programado(p: ReporteProgramado) -> dict:
    return {
        "id": str(p.id), "nombre": p.nombre, "tipo": p.tipo,
        "formato": p.formato, "params": p.params or {},
        "cron_expr": p.cron_expr, "activo": p.activo,
        "ultimo_run_at": _fecha_hora(p.ultimo_run_at) or None,
        "created_at": p.created_at.isoformat(),
    }


def _nombre_defecto(tipo: str) -> str:
    return f"{tipo.replace('_', ' ').capitalize()} {date.today().isoformat()}"


def _get_reporte(db: Session, cid: uuid.UUID, rid: uuid.UUID) -> Reporte:
    r = db.get(Reporte, rid)
    if not r or r.company_id != cid:
        raise err(404, "NO_ENCONTRADO", "El reporte no existe")
    return r


def _get_programado(db: Session, cid: uuid.UUID,
                    pid: uuid.UUID) -> ReporteProgramado:
    p = db.get(ReporteProgramado, pid)
    if not p or p.company_id != cid:
        raise err(404, "NO_ENCONTRADO", "La programación no existe")
    return p


# =============================================================================
# Historial y generación
# =============================================================================
@router.get("")
def listar(tipo: str | None = Query(None), formato: str | None = Query(None),
           status: str | None = Query(None),
           desde: date | None = Query(None), hasta: date | None = Query(None),
           p: Page = Depends(paginacion), db: Session = Depends(get_db),
           cid: uuid.UUID = Depends(get_company_id),
           user: User = Depends(get_current_user)):
    """Historial de reportes generados por la empresa, el más reciente primero.

    `generado_por` resuelve el usuario con un LEFT JOIN; cuando es NULL el
    reporte lo disparó una programación y se devuelve como «Sistema
    (Programado)», sin id ni email.
    """
    _solo_empresa(user)
    if tipo and tipo not in REPORTE_TIPOS:
        raise err(400, "TIPO_INVALIDO",
                  f"Tipo debe ser uno de: {', '.join(REPORTE_TIPOS)}")
    if formato:
        _valida_formato(formato, FORMATOS_REPORTE)
    if status and status not in ESTADOS_JOB:
        raise err(400, "STATUS_INVALIDO",
                  f"Status debe ser uno de: {', '.join(ESTADOS_JOB)}")
    if desde and hasta and hasta < desde:
        raise err(400, "RANGO_INVALIDO", "'hasta' no puede ser anterior a 'desde'")

    q = (select(Reporte, User.nombre, User.email)
         .select_from(Reporte)
         .outerjoin(User, User.id == Reporte.generado_por)
         .where(Reporte.company_id == cid))
    if tipo:
        q = q.where(Reporte.tipo == tipo)
    if formato:
        q = q.where(Reporte.formato == formato)
    if status:
        q = q.where(Reporte.status == status)
    if desde:
        q = q.where(Reporte.created_at >= datetime.combine(
            desde, datetime.min.time(), tzinfo=timezone.utc))
    if hasta:
        q = q.where(Reporte.created_at < datetime.combine(
            hasta + timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc))
    if p.search:
        q = q.where(Reporte.nombre.ilike(f"%{p.search}%"))

    total = db.scalar(select(func.count()).select_from(q.subquery())) or 0
    q = aplicar_orden(q, Reporte, p.sort, ORDEN_REPORTES, "-created_at")
    filas = db.execute(q.offset(p.offset).limit(p.page_size)).all()
    return sobre([_out(f[0], f.nombre, f.email) for f in filas], total, p)


@router.post("", status_code=202)
def crear(body: ReporteIn, db: Session = Depends(get_db),
          cid: uuid.UUID = Depends(get_company_id),
          user: User = Depends(require_company)):
    """Encola la generación de un reporte.

    Siempre asíncrono, incluso cuando el volumen es pequeño: el PDF y el Excel
    los produce el worker y así el tiempo de respuesta no depende del tamaño de
    la empresa. El resultado se consulta con `GET /reportes/{id}` y se baja con
    `GET /reportes/{id}/download-url`.
    """
    _solo_empresa(user)
    tipo = _valida_tipo(body.tipo)
    formato = _valida_formato(body.formato, FORMATOS_REPORTE)

    rep = Reporte(company_id=cid, nombre=(body.nombre or "").strip()
                  or _nombre_defecto(tipo), tipo=tipo, formato=formato,
                  status="queued", params=body.params or {},
                  generado_por=user.id)
    db.add(rep)
    db.flush()
    actividad.log(db, cid, "creacion", "reportes",
                  f"Reporte '{rep.nombre}' encolado ({tipo}, {formato})",
                  user_id=user.id, entidad_tipo="reporte", entidad_id=rep.id)
    db.commit()

    # Después del commit: con QUEUE_BACKEND=inproc la tarea corre en el acto y
    # abre su propia sesión, que solo ve lo ya confirmado.
    enqueue("generar_reporte", reporte_id=str(rep.id), company_id=str(cid))
    logger.info("reporte %s encolado (%s/%s)", rep.id, tipo, formato)
    return {"id": str(rep.id), "status": rep.status}


# --------------------------------------------------------------- programados
# Estas rutas se declaran ANTES de /{reporte_id}: si no, 'programados' entraría
# por el path param uuid y respondería 422 en vez de listar.
@router.get("/programados")
def listar_programados(activo: bool | None = Query(None),
                       tipo: str | None = Query(None),
                       p: Page = Depends(paginacion),
                       db: Session = Depends(get_db),
                       cid: uuid.UUID = Depends(get_company_id),
                       user: User = Depends(get_current_user)):
    """Programaciones de reportes recurrentes de la empresa."""
    _solo_empresa(user)
    q = select(ReporteProgramado).where(ReporteProgramado.company_id == cid)
    if activo is not None:
        q = q.where(ReporteProgramado.activo.is_(activo))
    if tipo:
        q = q.where(ReporteProgramado.tipo == tipo)
    if p.search:
        q = q.where(ReporteProgramado.nombre.ilike(f"%{p.search}%"))
    total = db.scalar(select(func.count()).select_from(q.subquery())) or 0
    q = aplicar_orden(q, ReporteProgramado, p.sort, ORDEN_PROGRAMADOS,
                      "-created_at")
    filas = list(db.scalars(q.offset(p.offset).limit(p.page_size)))
    return sobre([_out_programado(x) for x in filas], total, p)


@router.post("/programados", status_code=201)
def crear_programado(body: ProgramadoIn, db: Session = Depends(get_db),
                     cid: uuid.UUID = Depends(get_company_id),
                     user: User = Depends(require_company)):
    """Programa la generación recurrente de un reporte.

    Guardar la definición no dispara nada: la ejecución la agenda el beat del
    worker leyendo `cron_expr`. Por eso `activo=false` es una programación
    válida y no un error.
    """
    _solo_empresa(user)
    if not (body.nombre or "").strip():
        raise err(400, "NOMBRE_REQUERIDO", "El nombre es obligatorio")
    tipo = _valida_tipo(body.tipo)
    formato = _valida_formato(body.formato, FORMATOS_REPORTE)
    cron = valida_cron(body.cron_expr)

    p = ReporteProgramado(company_id=cid, nombre=body.nombre.strip(), tipo=tipo,
                          formato=formato, params=body.params or {},
                          cron_expr=cron, activo=body.activo)
    db.add(p)
    db.flush()
    actividad.log(db, cid, "creacion", "reportes",
                  f"Programación '{p.nombre}' creada ({cron}, {tipo}/{formato})",
                  user_id=user.id, entidad_tipo="reporte_programado",
                  entidad_id=p.id)
    db.commit()
    return _out_programado(p)


@router.patch("/programados/{programado_id}")
def editar_programado(programado_id: uuid.UUID, body: ProgramadoPatch,
                      db: Session = Depends(get_db),
                      cid: uuid.UUID = Depends(get_company_id),
                      user: User = Depends(require_company)):
    """Edita una programación: nombre, tipo, formato, params, cron o activo."""
    _solo_empresa(user)
    p = _get_programado(db, cid, programado_id)
    data = body.model_dump(exclude_unset=True)
    if not data:
        raise err(400, "SIN_CAMBIOS", "No se indicó ningún campo a actualizar")

    if "nombre" in data:
        if not (data["nombre"] or "").strip():
            raise err(400, "NOMBRE_REQUERIDO", "El nombre no puede estar vacío")
        p.nombre = data["nombre"].strip()
    if "tipo" in data and data["tipo"] is not None:
        p.tipo = _valida_tipo(data["tipo"])
    if "formato" in data and data["formato"] is not None:
        p.formato = _valida_formato(data["formato"], FORMATOS_REPORTE)
    if "params" in data and data["params"] is not None:
        p.params = data["params"]
    if "cron_expr" in data and data["cron_expr"] is not None:
        p.cron_expr = valida_cron(data["cron_expr"])
    if "activo" in data and data["activo"] is not None:
        p.activo = data["activo"]

    actividad.log(db, cid, "actualizacion", "reportes",
                  f"Programación '{p.nombre}' actualizada "
                  f"({', '.join(sorted(data))})", user_id=user.id,
                  entidad_tipo="reporte_programado", entidad_id=p.id)
    db.commit()
    return _out_programado(p)


@router.delete("/programados/{programado_id}")
def eliminar_programado(programado_id: uuid.UUID, db: Session = Depends(get_db),
                        cid: uuid.UUID = Depends(get_company_id),
                        user: User = Depends(require_company)):
    """Elimina una programación. Los reportes ya generados se conservan."""
    _solo_empresa(user)
    p = _get_programado(db, cid, programado_id)
    nombre = p.nombre
    db.delete(p)
    actividad.log(db, cid, "actualizacion", "reportes",
                  f"Programación '{nombre}' eliminada", user_id=user.id,
                  entidad_tipo="reporte_programado", entidad_id=programado_id)
    db.commit()
    return {"ok": True, "id": str(programado_id), "nombre": nombre,
            "reportes_conservados": True}


# ------------------------------------------------------------------- detalle
@router.get("/{reporte_id}")
def detalle(reporte_id: uuid.UUID, db: Session = Depends(get_db),
            cid: uuid.UUID = Depends(get_company_id),
            user: User = Depends(get_current_user)):
    """Estado del job y metadatos del reporte (endpoint de polling)."""
    _solo_empresa(user)
    r = _get_reporte(db, cid, reporte_id)
    autor = db.execute(select(User.nombre, User.email).where(
        User.id == r.generado_por)).first() if r.generado_por else None
    salida = _out(r, autor.nombre if autor else None,
                  autor.email if autor else None)
    salida["error"] = (r.params or {}).get("error")
    return salida


@router.get("/{reporte_id}/download-url")
def download_url(reporte_id: uuid.UUID, db: Session = Depends(get_db),
                 cid: uuid.UUID = Depends(get_company_id),
                 user: User = Depends(get_current_user)):
    """SAS de lectura del archivo generado.

    409 mientras el job no haya terminado: el archivo no existe todavía y una URL
    firmada apuntando a un blob ausente daría un 404 opaco en el navegador.
    """
    _solo_empresa(user)
    r = _get_reporte(db, cid, reporte_id)
    if r.status != "done" or not r.blob_path:
        raise err(409, "REPORTE_NO_LISTO",
                  f"El reporte está en estado '{r.status}'; "
                  "espere a que termine para descargarlo")
    ext = (r.blob_path.rsplit(".", 1)[-1] or "bin").lower()
    filename = nombre_archivo(r.nombre, r.formato, ext)
    dl = get_storage().download_url(r.blob_path, filename)
    return {"download_url": dl.download_url,
            "expires_at": dl.expires_at.isoformat(), "filename": filename}


# =============================================================================
# Exportaciones de vistas
# =============================================================================
@router_export.post("")
def exportar(body: ExportacionIn, db: Session = Depends(get_db),
             cid: uuid.UUID = Depends(get_company_id),
             user: User = Depends(require_company)):
    """Exporta la tabla en pantalla a Excel o CSV.

    Se cuentan primero las filas: hasta `EXPORT_FILAS_MAX` el archivo se genera
    en el propio request y se responde 200 con la URL de descarga; por encima se
    crea un `Reporte` equivalente, se encola y se responde 202. El umbral protege
    al worker de la API, que quedaría bloqueado escribiendo un XLSX de decenas de
    miles de filas mientras el navegador espera.

    Si `openpyxl` no está instalado el Excel degrada a CSV y el nombre del archivo
    lo dice (`...-excel-no-disponible.csv`), en vez de fallar la exportación.
    """
    _solo_empresa(user)
    recurso = body.recurso
    if recurso not in EXPORT_RECURSOS:
        raise err(400, "RECURSO_INVALIDO",
                  f"Recurso debe ser uno de: {', '.join(EXPORT_RECURSOS)}")
    if recurso not in CONSTRUCTORES:
        raise err(400, "RECURSO_NO_SOPORTADO",
                  f"El recurso '{recurso}' aún no tiene exportación implementada")
    formato = _valida_formato(body.formato, FORMATOS_EXPORT)
    scope = contrato_scope(user)
    filtros = body.filtros or {}

    n = contar_de_recurso(db, cid, recurso, filtros, scope)
    titulo = f"{recurso.capitalize()} {date.today().isoformat()}"

    if n > settings.export_filas_max:
        # Se deriva a un job. El recurso real va en params porque `reporte_tipo`
        # no tiene un valor por vista; la tarea lo prioriza sobre el tipo.
        rep = Reporte(
            company_id=cid, nombre=titulo,
            tipo=TIPO_POR_RECURSO.get(recurso, "estado_acreditacion"),
            formato="excel", status="queued",
            params={"recurso": recurso, "filtros": filtros,
                    "formato_export": formato},
            generado_por=user.id)
        db.add(rep)
        db.flush()
        actividad.log(db, cid, "creacion", "reportes",
                      f"Exportación de '{recurso}' derivada a reporte "
                      f"({n} filas superan el máximo de "
                      f"{settings.export_filas_max})", user_id=user.id,
                      entidad_tipo="reporte", entidad_id=rep.id)
        db.commit()
        enqueue("generar_reporte", reporte_id=str(rep.id), company_id=str(cid))
        logger.info("exportación de %s derivada a reporte %s (%d filas)",
                    recurso, rep.id, n)
        return JSONResponse(status_code=202, content={
            "id": str(rep.id), "status": rep.status, "filas_estimadas": n,
            "nota": ("La exportación excede el máximo síncrono; descárguela con "
                     "GET /reportes/{id}/download-url cuando el job termine."),
        })

    cabeceras, filas = filas_de_recurso(db, cid, recurso, filtros, scope)
    data, ext = serializar(formato, cabeceras, filas, titulo)
    filename = nombre_archivo(titulo, formato, ext)
    # Blob temporal: las exportaciones se purgan a las 24 horas (§10).
    blob_path = make_tmp_path(cid, f"export-{recurso}", filename)
    get_storage().save(blob_path, data)
    dl = get_storage().download_url(blob_path, filename)

    actividad.log(db, cid, "visualizacion", "reportes",
                  f"Exportación de '{recurso}' generada ({len(filas)} filas, "
                  f"{ext.upper()})", user_id=user.id)
    db.commit()
    logger.info("exportación de %s generada (%d filas, %s)", recurso,
                len(filas), ext)
    return {"download_url": dl.download_url,
            "expires_at": dl.expires_at.isoformat(),
            "filas": len(filas), "filename": filename,
            "formato": ext,
            "degradado_a_csv": bool(formato == "excel" and ext == "csv")}
